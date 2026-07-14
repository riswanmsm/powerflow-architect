import csv
import json
import os
from pathlib import Path
from typing import Dict, Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import Site

class Exporter:
    """
    Exports SharePoint inventory details to JSON, CSV, and Excel formats.
    Outputs are saved under the specified output directory (default: 'Inventory' at project root).
    """

    def __init__(self, output_dir: str = "Inventory"):
        self.output_dir = Path(output_dir)

    def export(self, site: Site) -> Dict[str, str]:
        """
        Export Site metadata, Lists, and Columns to JSON, CSV, and Excel.
        Returns a dictionary of generated file paths.
        """
        # Ensure output directory exists
        self.output_dir.mkdir(parents=True, exist_ok=True)

        json_path = self.output_dir / "inventory.json"
        csv_path = self.output_dir / "inventory.csv"
        xlsx_path = self.output_dir / "inventory.xlsx"

        self._export_json(site, json_path)
        self._export_csv(site, csv_path)
        self._export_xlsx(site, xlsx_path)

        return {
            "json": str(json_path),
            "csv": str(csv_path),
            "xlsx": str(xlsx_path),
        }

    def _export_json(self, site: Site, path: Path) -> None:
        """Export full hierarchical structure to JSON."""
        data = {
            "site_id": site.id,
            "site_name": site.name,
            "web_url": site.web_url,
            "hostname": site.hostname,
            "path": site.path,
            "lists": [
                {
                    "list_id": lst.id,
                    "list_name": lst.name,
                    "list_display_name": lst.display_name,
                    "web_url": lst.web_url,
                    "fields_count": len(lst.fields),
                    "fields": [
                        {
                            "field_id": fld.id,
                            "name": fld.name,
                            "display_name": fld.display_name,
                            "field_type": fld.field_type,
                            "is_required": fld.is_required,
                            "is_read_only": fld.is_read_only,
                            "is_hidden": fld.is_hidden,
                            "is_system": fld.is_system,
                        }
                        for fld in lst.fields
                    ]
                }
                for lst in site.lists
            ]
        }
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

    def _export_csv(self, site: Site, path: Path) -> None:
        """Export flattened columns data to CSV."""
        headers = [
            "List Name",
            "List ID",
            "Field Name",
            "Field Display Name",
            "Field ID",
            "Field Type",
            "Is Required",
            "Is Read Only",
            "Is Hidden",
            "Is System",
        ]

        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)

            for lst in site.lists:
                for fld in lst.fields:
                    writer.writerow([
                        lst.display_name,
                        lst.id,
                        fld.name,
                        fld.display_name,
                        fld.id,
                        fld.field_type,
                        str(fld.is_required),
                        str(fld.is_read_only),
                        str(fld.is_hidden),
                        str(fld.is_system),
                    ])

    def _export_xlsx(self, site: Site, path: Path) -> None:
        """Export styled, multi-sheet workbook to Excel with an Overview and Details tab."""
        wb = openpyxl.Workbook()
        
        # -------------------------------------------------------------
        # Tab 1: Overview Dashboard
        # -------------------------------------------------------------
        ws_overview = wb.active
        ws_overview.title = "Overview"
        ws_overview.views.sheetView[0].showGridLines = True

        # Styles
        font_title = Font(name="Segoe UI", size=16, bold=True, color="1B365D")
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        font_bold = Font(name="Segoe UI", size=11, bold=True)
        font_regular = Font(name="Segoe UI", size=11)
        
        fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        fill_stripe = PatternFill(start_color="F4F6F9", end_color="F4F6F9", fill_type="solid")
        fill_card = PatternFill(start_color="E9EDF0", end_color="E9EDF0", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )

        # Title Block
        ws_overview["A1"] = "SharePoint Site Inventory Overview"
        ws_overview["A1"].font = font_title
        ws_overview.row_dimensions[1].height = 30

        # Site Metadata Card
        ws_overview["A3"] = "Site Name:"
        ws_overview["A3"].font = font_bold
        ws_overview["B3"] = site.name
        ws_overview["B3"].font = font_regular

        ws_overview["A4"] = "Site URL:"
        ws_overview["A4"].font = font_bold
        ws_overview["B4"] = site.web_url
        ws_overview["B4"].font = font_regular

        ws_overview["A5"] = "Hostname:"
        ws_overview["A5"].font = font_bold
        ws_overview["B5"] = site.hostname
        ws_overview["B5"].font = font_regular

        # Metric Summaries
        total_lists = len(site.lists)
        total_fields = sum(len(lst.fields) for lst in site.lists)

        ws_overview["A7"] = "Total Lists:"
        ws_overview["A7"].font = font_bold
        ws_overview["B7"] = total_lists
        ws_overview["B7"].font = font_regular

        ws_overview["A8"] = "Total Fields:"
        ws_overview["A8"].font = font_bold
        ws_overview["B8"] = total_fields
        ws_overview["B8"].font = font_regular

        # Classification Breakdowns
        type_counts = {}
        for lst in site.lists:
            for fld in lst.fields:
                t = fld.field_type
                type_counts[t] = type_counts.get(t, 0) + 1

        # Write Classification counts table
        ws_overview["A11"] = "Field Type"
        ws_overview["A11"].font = font_header
        ws_overview["A11"].fill = fill_header
        ws_overview["A11"].alignment = Alignment(horizontal="left")

        ws_overview["B11"] = "Count"
        ws_overview["B11"].font = font_header
        ws_overview["B11"].fill = fill_header
        ws_overview["B11"].alignment = Alignment(horizontal="right")
        ws_overview.row_dimensions[11].height = 24

        # Sorted by count descending
        sorted_types = sorted(type_counts.items(), key=lambda x: x[1], reverse=True)
        
        current_row = 12
        for idx, (ftype, count) in enumerate(sorted_types):
            ws_overview.cell(row=current_row, column=1, value=ftype).font = font_regular
            c_cell = ws_overview.cell(row=current_row, column=2, value=count)
            c_cell.font = font_regular
            c_cell.alignment = Alignment(horizontal="right")
            
            # Apply thin borders & striping
            ws_overview.cell(row=current_row, column=1).border = thin_border
            c_cell.border = thin_border
            if idx % 2 == 1:
                ws_overview.cell(row=current_row, column=1).fill = fill_stripe
                c_cell.fill = fill_stripe
            
            ws_overview.row_dimensions[current_row].height = 20
            current_row += 1

        # -------------------------------------------------------------
        # Tab 2: Fields Inventory
        # -------------------------------------------------------------
        ws_details = wb.create_sheet(title="Fields Inventory")
        ws_details.views.sheetView[0].showGridLines = True

        headers = [
            "List Name",
            "List ID",
            "Field Name",
            "Field Display Name",
            "Field ID",
            "Field Type",
            "Is Required",
            "Is Read Only",
            "Is Hidden",
            "Is System",
        ]

        # Write Headers
        ws_details.row_dimensions[1].height = 26
        for col_idx, header in enumerate(headers, 1):
            cell = ws_details.cell(row=1, column=col_idx, value=header)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(vertical="center", horizontal="left")

        # Write Data
        row_idx = 2
        for lst in site.lists:
            for fld in lst.fields:
                ws_details.cell(row=row_idx, column=1, value=lst.display_name)
                ws_details.cell(row=row_idx, column=2, value=lst.id)
                ws_details.cell(row=row_idx, column=3, value=fld.name)
                ws_details.cell(row=row_idx, column=4, value=fld.display_name)
                ws_details.cell(row=row_idx, column=5, value=fld.id)
                ws_details.cell(row=row_idx, column=6, value=fld.field_type)
                
                # Use boolean format directly for easy Excel filtering/formulas
                ws_details.cell(row=row_idx, column=7, value=fld.is_required)
                ws_details.cell(row=row_idx, column=8, value=fld.is_read_only)
                ws_details.cell(row=row_idx, column=9, value=fld.is_hidden)
                ws_details.cell(row=row_idx, column=10, value=fld.is_system)

                # Styling cells
                for col_idx in range(1, 11):
                    c = ws_details.cell(row=row_idx, column=col_idx)
                    c.font = font_regular
                    c.border = thin_border
                    if row_idx % 2 == 1:
                        c.fill = fill_stripe

                ws_details.row_dimensions[row_idx].height = 20
                row_idx += 1

        # Enable Auto-filters on Details tab
        ws_details.auto_filter.ref = f"A1:J{row_idx - 1}"

        # Adjust Columns widths for both worksheets
        for ws in [ws_overview, ws_details]:
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                
                # Check lengths of values (avoid calculating on very long text to prevent slow execution)
                for cell in col:
                    val_str = str(cell.value or '')
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                
                # Limit size to prevent overly wide columns
                ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)

        # Save workbook
        wb.save(path)
