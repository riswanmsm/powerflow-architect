import json
from pathlib import Path
from typing import List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.template_engine.models import MappingTemplate

def clean_sheet_title(title: str) -> str:
    """Cleans a sheet title to prevent openpyxl errors (no special chars, max 31 chars)."""
    invalid_chars = ['\\', '/', '?', '*', '[', ']', ':']
    for char in invalid_chars:
        title = title.replace(char, '')
    return title[:31].strip() or "Sheet"

class JSONSerializer:
    """Serializes mapping templates to templates.json."""
    
    @staticmethod
    def serialize(templates: List[MappingTemplate], path: Path) -> None:
        data = {}
        for t in templates:
            data[t.list_name] = t.mappings
            
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

class ExcelSerializer:
    """Serializes mapping templates to templates.xlsx, optimized for copy-paste."""
    
    @staticmethod
    def serialize(templates: List[MappingTemplate], path: Path) -> None:
        wb = openpyxl.Workbook()
        # Remove default active sheet
        if wb.active:
            wb.remove(wb.active)

        # Styles
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        font_regular = Font(name="Segoe UI", size=11)
        fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        fill_stripe = PatternFill(start_color="F4F6F9", end_color="F4F6F9", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        align_left = Alignment(horizontal="left", vertical="center")

        headers = ["Excel Column Header", "Power Automate Expression"]

        for t in templates:
            sheet_title = clean_sheet_title(t.list_name)
            ws = wb.create_sheet(title=sheet_title)
            ws.views.sheetView[0].showGridLines = True
            
            # Write Headers
            ws.row_dimensions[1].height = 26
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_left

            # Write rows
            row_idx = 2
            for display_name, expr in t.mappings.items():
                ws.row_dimensions[row_idx].height = 20
                
                # Column 1: Excel Column Header
                c1 = ws.cell(row=row_idx, column=1, value=display_name)
                c1.font = font_regular
                c1.border = thin_border
                c1.alignment = align_left
                
                # Column 2: Power Automate Expression
                c2 = ws.cell(row=row_idx, column=2, value=expr)
                c2.font = font_regular
                c2.border = thin_border
                c2.alignment = align_left
                
                # Striping
                if row_idx % 2 == 1:
                    c1.fill = fill_stripe
                    c2.fill = fill_stripe
                    
                row_idx += 1

            # Auto-fit columns with safety margin
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or '')
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 120)

        wb.save(path)
