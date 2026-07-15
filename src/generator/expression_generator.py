import csv
import json
import os
from pathlib import Path
from typing import Dict, Any, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.powerautomate.expression import ExpressionContext, ExpressionEngine

def clean_sheet_title(title: str) -> str:
    """Cleans a sheet title to prevent openpyxl errors (no special chars, max 31 chars)."""
    invalid_chars = ['\\', '/', '?', '*', '[', ']', ':']
    for char in invalid_chars:
        title = title.replace(char, '')
    return title[:31].strip() or "Sheet"

class ExpressionGenerator:
    """
    Offline Expression Generator that reads a SharePoint list inventory JSON and
    generates copy-paste-ready Power Automate expressions in JSON, CSV, and Excel formats.
    """

    def __init__(self, input_json_path: str = "Inventory/inventory.json", output_dir: str = "output"):
        self.input_json_path = Path(input_json_path)
        self.output_dir = Path(output_dir)

    def generate(self, context: Optional[ExpressionContext] = None) -> Dict[str, Dict[str, str]]:
        """
        Loads the inventory metadata, generates the copy-paste-ready expressions,
        writes out JSON, CSV, and XLSX deliverables, and returns the nested expressions dict.
        """
        if context is None:
            context = ExpressionContext()

        if not self.input_json_path.exists():
            raise FileNotFoundError(f"Input inventory file not found at: {self.input_json_path}")

        # Ensure output directory exists
        self.output_dir.mkdir(parents=True, exist_ok=True)

        with open(self.input_json_path, "r", encoding="utf-8") as f:
            inventory = json.load(f)

        lists = inventory.get("lists", [])
        
        # 1. Compute nested expressions dictionary
        expressions_json_data = {}
        csv_rows = []
        excel_data = {}  # { sheet_title: [rows] }

        for lst in lists:
            list_name = lst.get("list_display_name") or lst.get("list_name") or "Unknown List"
            fields = lst.get("fields", [])
            
            list_exprs = {}
            excel_rows = []
            
            for field in fields:
                field_name = field.get("name") or ""
                is_system = field.get("is_system") is True
                if is_system and field_name.lower() != "title":
                    continue
                    
                display_name = field.get("display_name") or ""
                field_type_enum = ExpressionEngine.get_normalized_type(field)
                norm_type = field_type_enum.value
                
                # Generate expression
                expr = ExpressionEngine.generate(field, context)
                list_exprs[field_name] = expr
                
                # Append flat representations
                csv_rows.append([
                    list_name,
                    field_name,
                    display_name,
                    norm_type,
                    expr
                ])
                
                excel_rows.append([
                    field_name,
                    display_name,
                    norm_type,
                    expr
                ])

            expressions_json_data[list_name] = list_exprs
            excel_data[list_name] = excel_rows

        # 2. Write output/expressions.json
        json_path = self.output_dir / "expressions.json"
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(expressions_json_data, f, indent=2, ensure_ascii=False)

        # 3. Write output/expressions.csv
        csv_path = self.output_dir / "expressions.csv"
        csv_headers = [
            "List Name",
            "Internal Field Name",
            "Display Name",
            "Normalized Field Type",
            "Generated Expression"
        ]
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(csv_headers)
            writer.writerows(csv_rows)

        # 4. Write output/expressions.xlsx (Optimized for copy-paste)
        xlsx_path = self.output_dir / "expressions.xlsx"
        self._write_excel(excel_data, xlsx_path)

        return expressions_json_data

    def _write_excel(self, excel_data: dict, path: Path) -> None:
        wb = openpyxl.Workbook()
        # Remove default active sheet
        if wb.active:
            wb.remove(wb.active)

        # Styles
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        font_regular = Font(name="Segoe UI", size=11)
        fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        fill_stripe = PatternFill(start_color="F4F6F9", end_color="F4F6F9", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        align_left = Alignment(horizontal="left", vertical="center")

        headers = ["Internal Name", "Display Name", "Normalized Type", "Copy Value"]

        for list_name, rows in excel_data.items():
            sheet_title = clean_sheet_title(list_name)
            ws = wb.create_sheet(title=sheet_title)
            ws.views.sheetView[0].showGridLines = True
            
            # Write Headers
            ws.row_dimensions[1].height = 26
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_left

            # Write Rows
            row_idx = 2
            for r in rows:
                ws.row_dimensions[row_idx].height = 20
                for col_idx, val in enumerate(r, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = font_regular
                    cell.border = thin_border
                    cell.alignment = align_left
                    
                    # Apply striping
                    if row_idx % 2 == 1:
                        cell.fill = fill_stripe
                row_idx += 1

            # Auto-fit columns with safety margins
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or '')
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                # Keep columns neat but allow wide expressions column (column D / 4)
                ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 120)

        wb.save(path)
