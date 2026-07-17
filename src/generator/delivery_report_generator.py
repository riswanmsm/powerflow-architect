import csv
import json
import os
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.template_engine.template_engine import TemplateEngine

class DeliveryReportGenerator:
    """
    Delivery Report Generator compares expected flows for SharePoint lists
    against an externally supplied Power Automate flow inventory JSON.
    Generates reports in JSON, CSV, and Excel formats.
    """

    DEFAULT_FLOW_TEMPLATES = {
        "Manual Full Rebuild": "{list_name} - Manual Full Rebuild",
        "Scheduled Validation": "{list_name} - Scheduled Validation",
        "Delete Excel Row": "{list_name} - Delete Excel Row",
        "Sync Add and Update": "{list_name} to Excel Export - Sync Add and Update"
    }

    DEFAULT_SYSTEM_LISTS = {
        "Documents",
        "Form Templates",
        "Web Template Extensions",
        "Site Assets",
        "Style Library",
        "Site Pages",
        "Master Page Gallery"
    }

    def __init__(
        self,
        inventory_path: str = "Inventory/inventory.json",
        templates_path: str = "output/templates.json",
        existing_flows_path: Optional[str] = None
    ):
        self.inventory_path = Path(inventory_path)
        self.templates_path = Path(templates_path)
        self.existing_flows_path = Path(existing_flows_path) if existing_flows_path else None

    def generate(self, output_dir: str = "output") -> Dict[str, Any]:
        """
        Generates the delivery report files (JSON, CSV, XLSX) and returns a summary dict.
        """
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)

        # 1. Load SharePoint list names reusing TemplateEngine to preserve list filtering/discovery
        if not self.inventory_path.exists():
            raise FileNotFoundError(f"Inventory file not found at: {self.inventory_path}")
        
        temp_engine = TemplateEngine(inventory_path=str(self.inventory_path))
        mapping_templates = temp_engine.load_templates()
        list_names = [m.list_name for m in mapping_templates]

        # If list_names is empty, let's fallback to inventory.json lists directly to ensure robust handling
        if not list_names:
            with open(self.inventory_path, "r", encoding="utf-8") as f:
                inv_data = json.load(f)
            lists_data = inv_data.get("lists", [])
            list_names = [
                lst.get("list_display_name") or lst.get("list_name") or "Unknown List"
                for lst in lists_data
            ]

        # Remove duplicates while preserving order
        seen = set()
        list_names = [x for x in list_names if not (x in seen or seen.add(x))]

        # Filter out library lists starting with LIB and default system lists
        list_names = [
            name for name in list_names
            if not name.startswith("LIB") and name not in self.DEFAULT_SYSTEM_LISTS
        ]

        # 2. Load existing flows inventory if available
        existing_flows_lookup = {}
        if self.existing_flows_path and self.existing_flows_path.exists():
            try:
                with open(self.existing_flows_path, "r", encoding="utf-8") as f:
                    flows_inventory = json.load(f)
                
                # Check if it's a list (as expected)
                if isinstance(flows_inventory, list):
                    for flow in flows_inventory:
                        properties = flow.get("properties", {})
                        display_name = properties.get("displayName")
                        if display_name:
                            existing_flows_lookup[display_name] = {
                                "id": flow.get("name") or flow.get("id", ""),
                                "state": properties.get("state") or "",
                                "createdTime": properties.get("createdTime") or "",
                                "lastModifiedTime": properties.get("lastModifiedTime") or ""
                            }
            except Exception as e:
                print(f"Warning: Failed to parse existing_flows.json: {e}. Proceeding without external flow comparison.")

        # 3. Determine expected flows and match status
        report_rows = []
        total_expected = 0
        total_complete = 0
        total_missing = 0

        for list_name in list_names:
            for flow_type, pattern in self.DEFAULT_FLOW_TEMPLATES.items():
                expected_name = pattern.format(list_name=list_name)
                flow_info = existing_flows_lookup.get(expected_name)

                exists = "Yes" if flow_info else "No"
                status = "Complete" if flow_info else "Missing"

                if flow_info:
                    total_complete += 1
                else:
                    total_missing += 1
                total_expected += 1

                report_rows.append({
                    "ListName": list_name,
                    "ExpectedFlowName": expected_name,
                    "FlowType": flow_type,
                    "Exists": exists,
                    "Status": status,
                    "FlowID": flow_info["id"] if flow_info else "",
                    "FlowState": flow_info["state"] if flow_info else "",
                    "CreatedTime": flow_info["createdTime"] if flow_info else "",
                    "LastModifiedTime": flow_info["lastModifiedTime"] if flow_info else ""
                })

        summary = {
            "total_lists_evaluated": len(list_names),
            "total_expected_flows": total_expected,
            "completed_flows_count": total_complete,
            "missing_flows_count": total_missing,
            "comparison_active": bool(existing_flows_lookup)
        }

        # 4. Write delivery_report.json
        json_data = {
            "generated_at": datetime.now().isoformat(),
            "summary": summary,
            "flows": report_rows
        }
        json_path = output_path / "delivery_report.json"
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(json_data, f, indent=2, ensure_ascii=False)

        # 5. Write delivery_report.csv
        csv_path = output_path / "delivery_report.csv"
        csv_headers = [
            "List Name",
            "Expected Flow Name",
            "Flow Type",
            "Exists",
            "Status",
            "Flow ID",
            "Flow State",
            "Created Time",
            "Last Modified Time"
        ]
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(csv_headers)
            for row in report_rows:
                writer.writerow([
                    row["ListName"],
                    row["ExpectedFlowName"],
                    row["FlowType"],
                    row["Exists"],
                    row["Status"],
                    row["FlowID"],
                    row["FlowState"],
                    row["CreatedTime"],
                    row["LastModifiedTime"]
                ])

        # 6. Write delivery_report.xlsx (styled/premium)
        xlsx_path = output_path / "delivery_report.xlsx"
        self._write_excel(report_rows, xlsx_path)

        return summary

    def _write_excel(self, report_rows: List[Dict[str, Any]], path: Path) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Delivery Report"
        ws.views.sheetView[0].showGridLines = True

        # Styles
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        font_regular = Font(name="Segoe UI", size=11)
        font_complete = Font(name="Segoe UI", size=11, bold=True, color="155724")
        font_missing = Font(name="Segoe UI", size=11, bold=True, color="721C24")

        fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        fill_stripe = PatternFill(start_color="F4F6F9", end_color="F4F6F9", fill_type="solid")
        fill_complete = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        fill_missing = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )

        align_left = Alignment(horizontal="left", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")

        headers = [
            "List Name",
            "Expected Flow Name",
            "Flow Type",
            "Exists",
            "Status",
            "Flow ID",
            "Flow State",
            "Created Time",
            "Last Modified Time"
        ]

        # Write headers
        ws.row_dimensions[1].height = 26
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_left

        # Write data rows
        row_idx = 2
        for r in report_rows:
            ws.row_dimensions[row_idx].height = 20
            
            # Map standard values to cells
            vals = [
                r["ListName"],
                r["ExpectedFlowName"],
                r["FlowType"],
                r["Exists"],
                r["Status"],
                r["FlowID"],
                r["FlowState"],
                r["CreatedTime"],
                r["LastModifiedTime"]
            ]

            for col_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = font_regular
                cell.border = thin_border
                cell.alignment = align_left

                # Align centered columns
                if headers[col_idx - 1] in ["Exists", "Status", "Flow State", "Created Time", "Last Modified Time"]:
                    cell.alignment = align_center

                # Apply conditional status styling or general striping
                if headers[col_idx - 1] == "Exists":
                    if val == "Yes":
                        cell.fill = fill_complete
                        cell.font = font_complete
                    else:
                        cell.fill = fill_missing
                        cell.font = font_missing
                elif headers[col_idx - 1] == "Status":
                    if val == "Complete":
                        cell.fill = fill_complete
                        cell.font = font_complete
                    else:
                        cell.fill = fill_missing
                        cell.font = font_missing
                else:
                    if row_idx % 2 == 1:
                        cell.fill = fill_stripe

            row_idx += 1

        # Enable Auto-filters on the entire range
        ws.auto_filter.ref = f"A1:I{row_idx - 1}"

        # Adjust Columns widths for worksheets auto-fit
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 60)

        wb.save(path)
