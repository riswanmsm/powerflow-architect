import json
import pytest
import openpyxl
from pathlib import Path

from src.powerautomate.expression import ExpressionContext
from src.template_engine.template_engine import TemplateEngine

@pytest.fixture
def mock_inventory_for_templates(tmp_path):
    inventory_data = {
        "site_id": "site-abc",
        "site_name": "Sync Site",
        "lists": [
            {
                "list_id": "lst-tasks",
                "list_name": "tasks",
                "list_display_name": "Tasks List",
                "fields": [
                    {
                        "field_id": "f1",
                        "name": "Title",
                        "display_name": "Task Title",
                        "is_system": True,  # Title is system, but must be kept
                        "field_type": "Text"
                    },
                    {
                        "field_id": "f2",
                        "name": "ID",
                        "display_name": "ID",
                        "is_system": True,  # ID is system, must be excluded
                        "field_type": "Counter"
                    },
                    {
                        "field_id": "f3",
                        "name": "Status",
                        "display_name": "Current Status",
                        "is_system": False,
                        "field_type": "Choice"
                    },
                    {
                        "field_id": "f4",
                        "name": "Geoloc",
                        "display_name": "Geolocation",
                        "is_system": False,
                        "field_type": "Geolocation"  # Resolves to UNKNOWN, must be excluded
                    }
                ]
            }
        ]
    }
    input_path = tmp_path / "inventory.json"
    with open(input_path, "w", encoding="utf-8") as f:
        json.dump(inventory_data, f)
    return input_path

def test_template_engine_loads_and_filters_correctly(mock_inventory_for_templates):
    engine = TemplateEngine(inventory_path=str(mock_inventory_for_templates))
    templates = engine.load_templates()
    
    assert len(templates) == 1
    t = templates[0]
    
    assert t.list_name == "Tasks List"
    
    # Verify filtering:
    # 1. 'Task Title' should be present (Title is kept despite is_system=True)
    assert "Task Title" in t.mappings
    assert t.mappings["Task Title"] == "@items('Apply_to_each_1')?['Title']"
    
    # 2. 'ID' should NOT be present (read-only system field)
    assert "ID" not in t.mappings
    
    # 3. 'Current Status' should be present (standard choice field)
    assert "Current Status" in t.mappings
    assert t.mappings["Current Status"] == "@items('Apply_to_each_1')?['Status/Value']"
    
    # 4. 'Geolocation' should NOT be present (unsupported/UNKNOWN field type)
    assert "Geolocation" not in t.mappings

def test_template_engine_generates_correct_files(tmp_path, mock_inventory_for_templates):
    output_dir = tmp_path / "output"
    engine = TemplateEngine(inventory_path=str(mock_inventory_for_templates))
    
    templates = engine.generate_outputs(output_dir=str(output_dir))
    
    json_out = output_dir / "templates.json"
    xlsx_out = output_dir / "templates.xlsx"
    
    assert json_out.exists()
    assert xlsx_out.exists()
    
    # 1. Verify JSON file contents
    with open(json_out, "r", encoding="utf-8") as f:
        data = json.load(f)
    assert data == {"Tasks List": templates[0].mappings}
    
    # 2. Verify Excel workbook structure & headers
    wb = openpyxl.load_workbook(xlsx_out)
    assert "Tasks List" in wb.sheetnames
    
    ws = wb["Tasks List"]
    assert ws.views.sheetView[0].showGridLines is True
    
    # Headers
    assert ws.cell(row=1, column=1).value == "Excel Column Header"
    assert ws.cell(row=1, column=2).value == "Power Automate Expression"
    
    # Rows (Task Title & Current Status)
    assert ws.cell(row=2, column=1).value == "Task Title"
    assert ws.cell(row=2, column=2).value == "@items('Apply_to_each_1')?['Title']"
    
    assert ws.cell(row=3, column=1).value == "Current Status"
    assert ws.cell(row=3, column=2).value == "@items('Apply_to_each_1')?['Status/Value']"

def test_template_engine_with_custom_context(tmp_path, mock_inventory_for_templates):
    output_dir = tmp_path / "output"
    engine = TemplateEngine(inventory_path=str(mock_inventory_for_templates))
    
    custom_context = ExpressionContext(loop_name="For_each_row")
    templates = engine.generate_outputs(output_dir=str(output_dir), context=custom_context)
    
    assert templates[0].mappings["Task Title"] == "@items('For_each_row')?['Title']"
