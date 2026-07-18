import csv
import json
import openpyxl
import pytest
from pathlib import Path
from src.generator.expression_generator import ExpressionGenerator
from src.powerautomate.expression import ExpressionContext

@pytest.fixture
def mock_inventory_file(tmp_path):
    inventory_data = {
        "site_id": "site-xyz",
        "site_name": "Test Site",
        "lists": [
            {
                "list_id": "lst-1",
                "list_name": "projects_list",
                "list_display_name": "Projects",
                "fields": [
                    {
                        "field_id": "f1",
                        "name": "Title",
                        "display_name": "Project Name",
                        "NormalizedFieldType": "Text"
                    },
                    {
                        "field_id": "f2",
                        "name": "Tags",
                        "display_name": "Project Tags",
                        "NormalizedFieldType": "ChoiceMulti"
                    },
                    {
                        "field_id": "f3",
                        "name": "RelatedTask",
                        "display_name": "Related Task",
                        "NormalizedFieldType": "Lookup"
                    }
                ]
            }
        ]
    }
    input_path = tmp_path / "mock_inventory.json"
    with open(input_path, "w", encoding="utf-8") as f:
        json.dump(inventory_data, f)
    return input_path

def test_generator_runs_offline_and_creates_all_outputs(tmp_path, mock_inventory_file):
    output_dir = tmp_path / "output"
    
    # Initialize generator
    generator = ExpressionGenerator(
        input_json_path=str(mock_inventory_file),
        output_dir=str(output_dir)
    )
    
    # Generate standard context expressions
    res = generator.generate()
    
    # Verify return dict
    assert "Projects" in res
    assert "Title" in res["Projects"]
    assert res["Projects"]["Title"]["direct_trigger"] == "@triggerBody()?['Title']"
    assert res["Projects"]["Title"]["foreach"] == "@items('Apply_to_each_1')?['Title']"
    assert "RelatedTask" in res["Projects"]
    assert res["Projects"]["RelatedTask"]["direct_trigger"] == "@triggerBody()?['RelatedTask/Value']"
    assert res["Projects"]["RelatedTask"]["foreach"] == "@items('Apply_to_each_1')?['RelatedTask/Value']"
    
    # Verify outputs exist
    json_out = output_dir / "expressions.json"
    csv_out = output_dir / "expressions.csv"
    xlsx_out = output_dir / "expressions.xlsx"
    
    assert json_out.exists()
    assert csv_out.exists()
    assert xlsx_out.exists()
    
    # 1. Verify JSON contents
    with open(json_out, "r", encoding="utf-8") as f:
        loaded_json = json.load(f)
    assert loaded_json == res
    
    # 2. Verify CSV contents
    expected_headers = [
        "List Name",
        "Internal Field Name",
        "Display Name",
        "Normalized Field Type",
        "Direct Trigger Expression",
        "Foreach Expression"
    ]
    with open(csv_out, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        headers = next(reader)
        assert headers == expected_headers
        
        rows = list(reader)
        assert len(rows) == 3
        
        # Verify Title row
        assert rows[0][0] == "Projects"
        assert rows[0][1] == "Title"
        assert rows[0][2] == "Project Name"
        assert rows[0][3] == "Text"
        assert rows[0][4] == "@triggerBody()?['Title']"
        assert rows[0][5] == "@items('Apply_to_each_1')?['Title']"
        
        # Verify ChoiceMulti row (xpath structure)
        assert rows[1][0] == "Projects"
        assert rows[1][1] == "Tags"
        assert rows[1][3] == "ChoiceMulti"
        assert "xpath" in rows[1][4]
        assert "xpath" in rows[1][5]
        
    # 3. Verify Excel workbook worksheets & structure
    wb = openpyxl.load_workbook(xlsx_out)
    assert "Projects" in wb.sheetnames
    
    ws = wb["Projects"]
    # Check gridlines setting
    assert ws.views.sheetView[0].showGridLines is True
    
    # Check headers
    assert ws.cell(row=1, column=1).value == "Internal Name"
    assert ws.cell(row=1, column=2).value == "Display Name"
    assert ws.cell(row=1, column=3).value == "Normalized Type"
    assert ws.cell(row=1, column=4).value == "Direct Trigger Expression"
    assert ws.cell(row=1, column=5).value == "Foreach Expression"
    
    # Check Title row cells
    assert ws.cell(row=2, column=1).value == "Title"
    assert ws.cell(row=2, column=2).value == "Project Name"
    assert ws.cell(row=2, column=3).value == "Text"
    assert ws.cell(row=2, column=4).value == "@triggerBody()?['Title']"
    assert ws.cell(row=2, column=5).value == "@items('Apply_to_each_1')?['Title']"

def test_generator_with_custom_context(tmp_path, mock_inventory_file):
    output_dir = tmp_path / "output"
    generator = ExpressionGenerator(
        input_json_path=str(mock_inventory_file),
        output_dir=str(output_dir)
    )
    
    custom_context = ExpressionContext(
        loop_name="For_each_item",
        delimiter=" ; ",
        lookup_property="Title"
    )
    
    res = generator.generate(custom_context)
    
    # Verify Title
    assert res["Projects"]["Title"]["direct_trigger"] == "@triggerBody()?['Title']"
    assert res["Projects"]["Title"]["foreach"] == "@items('For_each_item')?['Title']"
    
    # Verify Lookup uses the custom lookup_property Title
    assert res["Projects"]["RelatedTask"]["direct_trigger"] == "@triggerBody()?['RelatedTask/Title']"
    assert res["Projects"]["RelatedTask"]["foreach"] == "@items('For_each_item')?['RelatedTask/Title']"
