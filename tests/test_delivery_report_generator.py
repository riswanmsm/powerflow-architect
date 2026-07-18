import csv
import json
import openpyxl
import pytest
from pathlib import Path
from src.generator.delivery_report_generator import DeliveryReportGenerator

@pytest.fixture
def mock_inventory(tmp_path):
    inventory_data = {
        "site_id": "site-id-123",
        "site_name": "Test Site",
        "lists": [
            {
                "list_id": "lst-1",
                "list_name": "REG_Risks",
                "list_display_name": "REG_Risks",
                "fields": [
                    {
                        "field_id": "f1",
                        "name": "Title",
                        "display_name": "Title",
                        "is_system": True,
                        "field_type": "Text"
                    }
                ]
            },
            {
                "list_id": "lst-2",
                "list_name": "LIB_Docs",
                "list_display_name": "LIB_Docs",
                "fields": [
                    {
                        "field_id": "f2",
                        "name": "Title",
                        "display_name": "Title",
                        "is_system": True,
                        "field_type": "Text"
                    }
                ]
            },
            {
                "list_id": "lst-3",
                "list_name": "Web Template Extensions",
                "list_display_name": "Web Template Extensions",
                "fields": [
                    {
                        "field_id": "f3",
                        "name": "Title",
                        "display_name": "Title",
                        "is_system": True,
                        "field_type": "Text"
                    }
                ]
            },
            {
                "list_id": "lst-4",
                "list_name": "Documents",
                "list_display_name": "Documents",
                "fields": [
                    {
                        "field_id": "f4",
                        "name": "Title",
                        "display_name": "Title",
                        "is_system": True,
                        "field_type": "Text"
                    }
                ]
            }
        ]
    }
    path = tmp_path / "mock_inventory.json"
    with open(path, "w", encoding="utf-8") as f:
        json.dump(inventory_data, f)
    return path

@pytest.fixture
def mock_templates(tmp_path):
    templates_data = {
        "REG_Risks": {
            "Title": "@items('Apply_to_each_1')?['Title']"
        },
        "LIB_Docs": {
            "Title": "@items('Apply_to_each_1')?['Title']"
        },
        "Web Template Extensions": {
            "Title": "@items('Apply_to_each_1')?['Title']"
        },
        "Documents": {
            "Title": "@items('Apply_to_each_1')?['Title']"
        }
    }
    path = tmp_path / "mock_templates.json"
    with open(path, "w", encoding="utf-8") as f:
        json.dump(templates_data, f)
    return path

@pytest.fixture
def mock_existing_flows(tmp_path):
    flows_data = [
        {
            "name": "flow-uuid-1",
            "id": "/flows/flow-uuid-1",
            "properties": {
                "displayName": "REG_Risks - Manual Full Rebuild",
                "state": "Started",
                "createdTime": "2026-07-13T12:00:00Z",
                "lastModifiedTime": "2026-07-13T14:00:00Z"
            }
        },
        {
            "name": "flow-uuid-2",
            "id": "/flows/flow-uuid-2",
            "properties": {
                "displayName": "REG_Risks to Excel Export - Sync Add and Update",
                "state": "Stopped",
                "createdTime": "2026-07-14T09:00:00Z",
                "lastModifiedTime": "2026-07-14T10:00:00Z"
            }
        }
    ]
    path = tmp_path / "mock_existing_flows.json"
    with open(path, "w", encoding="utf-8") as f:
        json.dump(flows_data, f)
    return path

def test_delivery_report_generator_omitted_existing_flows(tmp_path, mock_inventory, mock_templates):
    output_dir = tmp_path / "output"
    
    generator = DeliveryReportGenerator(
        inventory_path=str(mock_inventory),
        templates_path=str(mock_templates),
        existing_flows_path=None
    )
    
    summary = generator.generate(output_dir=str(output_dir))
    
    # 1 list (excluding LIB_Docs) * 4 flow types = 4 expected flows
    assert summary["total_lists_evaluated"] == 1
    assert summary["total_expected_flows"] == 4
    assert summary["completed_flows_count"] == 0
    assert summary["missing_flows_count"] == 4
    assert summary["comparison_active"] is False

    # Check that outputs are created
    json_report = output_dir / "delivery_report.json"
    csv_report = output_dir / "delivery_report.csv"
    xlsx_report = output_dir / "delivery_report.xlsx"
    
    assert json_report.exists()
    assert csv_report.exists()
    assert xlsx_report.exists()

    # Load JSON and verify
    with open(json_report, "r", encoding="utf-8") as f:
        data = json.load(f)
    
    assert len(data["flows"]) == 4
    for row in data["flows"]:
        assert row["Exists"] == "No"
        assert row["Status"] == "Missing"
        assert row["FlowID"] == ""

    # Load CSV and verify headers
    with open(csv_report, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        headers = next(reader)
        assert headers == [
            "List Name", "Expected Flow Name", "Flow Type",
            "Exists", "Status", "Flow ID", "Flow State",
            "Created Time", "Last Modified Time"
        ]

def test_delivery_report_generator_with_existing_flows(tmp_path, mock_inventory, mock_templates, mock_existing_flows):
    output_dir = tmp_path / "output"
    
    generator = DeliveryReportGenerator(
        inventory_path=str(mock_inventory),
        templates_path=str(mock_templates),
        existing_flows_path=str(mock_existing_flows)
    )
    
    summary = generator.generate(output_dir=str(output_dir))
    
    # Total expected: 4, Completed: 2, Missing: 2
    assert summary["total_expected_flows"] == 4
    assert summary["completed_flows_count"] == 2
    assert summary["missing_flows_count"] == 2
    assert summary["comparison_active"] is True

    json_report = output_dir / "delivery_report.json"
    with open(json_report, "r", encoding="utf-8") as f:
        data = json.load(f)
        
    flows = data["flows"]
    # Check that "REG_Risks - Manual Full Rebuild" is marked complete
    risks_rebuild = next(x for x in flows if x["ExpectedFlowName"] == "REG_Risks - Manual Full Rebuild")
    assert risks_rebuild["Exists"] == "Yes"
    assert risks_rebuild["Status"] == "Complete"
    assert risks_rebuild["FlowID"] == "flow-uuid-1"
    assert risks_rebuild["FlowState"] == "Started"
    assert risks_rebuild["CreatedTime"] == "2026-07-13T12:00:00Z"

    # Check that "REG_Risks to Excel Export - Sync Add and Update" is marked complete
    risks_sync = next(x for x in flows if x["ExpectedFlowName"] == "REG_Risks to Excel Export - Sync Add and Update")
    assert risks_sync["Exists"] == "Yes"
    assert risks_sync["Status"] == "Complete"
    assert risks_sync["FlowID"] == "flow-uuid-2"
    assert risks_sync["FlowState"] == "Stopped"

    # Check a missing one
    risks_delete = next(x for x in flows if x["ExpectedFlowName"] == "REG_Risks - Delete Excel Row")
    assert risks_delete["Exists"] == "No"
    assert risks_delete["Status"] == "Missing"
    assert risks_delete["FlowID"] == ""

    # Verify Excel styling/loading
    wb = openpyxl.load_workbook(output_dir / "delivery_report.xlsx")
    assert "Delivery Report" in wb.sheetnames
    ws = wb["Delivery Report"]
    
    # Header check
    assert ws.cell(row=1, column=1).value == "List Name"
    assert ws.cell(row=1, column=4).value == "Exists"
    
    # Check styling for exists complete
    # Find complete row for REG_Risks - Manual Full Rebuild
    complete_row = None
    for r in range(2, 10):
        if ws.cell(row=r, column=2).value == "REG_Risks - Manual Full Rebuild":
            complete_row = r
            break
            
    assert complete_row is not None
    exists_cell = ws.cell(row=complete_row, column=4)
    assert exists_cell.value == "Yes"
    # Complete text is color "155724", fill "D4EDDA"
    assert exists_cell.font.color.rgb == "00155724" or exists_cell.font.color.value == "155724"
    assert exists_cell.fill.start_color.rgb == "00D4EDDA" or exists_cell.fill.start_color.value == "D4EDDA"
